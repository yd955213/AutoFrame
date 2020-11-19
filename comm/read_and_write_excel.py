# !/uer/bin/env python
# -*- coding: utf-8 -*-
"""
@File   :   read_and_write_excel.py
@Time   :   2020-11-12 9:32
@Author :   yang_dang
@Contact    :   664720125@qq.com
@Version    :   1.0
@Description   :  实现excel的读取，支持xls,xlxs
"""

import os
import shutil
from enum import Enum
import openpyxl
import xlrd
import xlwt
from openpyxl.styles import PatternFill, Font
from xlutils.copy import copy
from global_abspath import get_abspath

# """
# xlrd\xlwt、openpyxl的背景颜色顺序需一一对应 7之后的有空再写
# 背景色 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon,
# 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal,
# 22 = Light Gray, 23 = Dark Gray
# """


class MyColor(Enum):
    BlACK = '000000'
    WHITE = 'FFFFFF'
    RED = 'FF0000'
    GREEN = '00FF00'
    BLUE = '0000FF'
    YELLOW = 'FFFF00'
    MAGENTA = 'FF00FF'
    CYAN = '00FFFF'


class Excel:
    """
    实现Excel文档的读写操作，
    xlrd、xlwt操作Excel时，下标从0开始，
    openpyxl操作Excel时，下标从1开始；
    为同时支持，规定下标从0开始，openpyxl获取的小标进行加1操作
    """
    excel_type_xls = 'xls'
    excel_type_xlsx = 'xlsx'

    def __init__(self):
        # 读取需要复制的excel的工作空间
        self.workbook = None
        # 写需要复制的excel的工作空间
        self.workbook_write = None
        # 当前工作的sheet页
        self.sheet = None
        # 当前工作的sheet页
        self.sheet_write = None
        # 当前sheet的行数
        self.rows = 0
        # 当前读取到的行数
        self.readingLine = 0
        # excel 类型：xls，xlsx
        self.excel_type = None
        # excel保存路径
        self.result_file = None

    def open_excel(self, file_name):
        file_path = get_abspath(file_name)
        if file_path is None:
            return
        else:
            self.result_file = os.path.dirname(file_path) + os.sep + 'result_' + os.path.basename(file_name)

        if file_path.endswith(self.excel_type_xls):
            xlrd.Book.encoding = 'utf8'
            # formatting_info带格式的复制
            self.workbook = xlrd.open_workbook(filename=file_path, formatting_info=True)
            # 拷贝，也在内存里面
            shutil.copy(file_path, self.result_file)
            # 打开备份文件进行读写操作
            self.workbook_write = copy(self.workbook)
            self.sheet = self.workbook.sheet_by_index(0)
            self.sheet_write = self.sheet
            self.rows = self.sheet.nrows
            self.excel_type = self.excel_type_xls
        else:
            openpyxl.Workbook.encoding = 'utf8'
            self.workbook = openpyxl.load_workbook(filename=file_path)
            shutil.copy(file_path, self.result_file)
            self.workbook_write = openpyxl.load_workbook(filename=self.result_file)
            self.sheet = self.workbook[self.workbook.sheetnames[0]]
            self.rows = self.sheet.max_row
            self.excel_type = self.excel_type_xlsx

        self.readingLine = 0

    def get_sheets(self):
        if self.excel_type == self.excel_type_xls:
            return self.__get_sheets_xls()
        else:
            return self.__get_sheets_xlsx()

    def set_sheet(self, name):
        if self.excel_type == self.excel_type_xls:
            self.__set_sheet_xls(name)
        else:
            self.__set_sheet_xlsx(name)

    def read_line(self):
        if self.excel_type == self.excel_type_xls:
            return self.__read_line_xls()
        else:
            return self.__read_line_xlsx()

    def save(self):
        self.workbook_write.save(self.result_file)

    def write(self, row, column, value, color=MyColor.WHITE):
        if self.excel_type == self.excel_type_xls:
            self.__write_xls(row, column, value, color)
        else:
            self.__write_xlsx(row, column, value, color)

    def __get_sheets_xls(self):
        return self.workbook.sheet_names()

    def __get_sheets_xlsx(self):
        return self.workbook.sheetnames

    def __set_sheet_xls(self, name):
        self.sheet = self.workbook.sheet_by_name(name)
        self.sheet_write = self.workbook_write.get_sheet(name)
        self.rows = self.sheet.nrows
        self.readingLine = 0
        return

    def __set_sheet_xlsx(self, name):
        self.sheet = self.workbook_write[name]
        self.sheet_write = self.sheet
        self.rows = self.sheet.max_row
        self.readingLine = 0
        return

    def __read_line_xls(self):
        lines = []
        if self.readingLine < self.rows:
            row = self.sheet.row_values(self.readingLine)
            self.readingLine += 1
            # row_str = []
            for value in row:
                lines.append(str(value))
        return lines

    def __read_line_xlsx(self):
        lines = []
        for row in self.sheet.rows:
            value = []
            for cell in row:
                if cell.value is None:
                    value.append('')
                else:
                    value.append(cell.value)
            lines.append(value)
        return lines

    def __write_xls(self, row, column, value, color=None):
        # 获取单元格格式
        # cell_style = self.sheet.cell(row, column).xf_idx
        # self.sheet.write(row=row, column=column, value=value)
        # 获取要写入的单元格
        # 获取要写入的单元格
        def _getCell(sheet, r, c):
            """ HACK: Extract the internal xlwt cell representation. """
            # 获取行
            row = sheet._Worksheet__rows.get(r)
            if not row:
                return None

            # 获取单元格
            cell = row._Row__cells.get(c)
            return cell

        row = int(row)
        column = int(column)
        if row >= 0 and column >= 0:
            # 获取要写入的单元格
            cell = _getCell(self.sheet_write, row, column)
            # 格式，把单元格原来的格式保存下来
            idx = cell.xf_idx

            if color is None:
                self.sheet_write.write(row, column, value)
                if cell:
                    # 获取要写入的单元格
                    ncell = _getCell(self.sheet_write, row, column)
                    if ncell:
                        # 设置写入后格式和写入前一样
                        ncell.xf_idx = idx
            else:
                # 初始化样式
                style = xlwt.XFStyle()

                font1 = xlwt.Font()  # 创建字体
                font1.name = 'Arial'
                font1.bold = True  # 黑体
                # font.underline = True  # 下划线
                # font.italic = True  # 斜体字
                font1.colour_index = 0  # 颜色为红色
                style.font = font1
                # borders.left = xlwt.Borders.THIN
                # DASHED 虚线
                # NO_LINE： 官方代码中NO_LINE所表示的值为0，没有边框
                # THIN： 官方代码中THIN所表示的值为1，边框为实线
                border = xlwt.Borders()         # 创建边框
                border.left = xlwt.Borders.THIN
                border.right = xlwt.Borders.THIN
                border.top = xlwt.Borders.THIN
                border.bottom = xlwt.Borders.THIN
                style.borders = border
                '''
                背景色 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon,
                17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal,
                22 = Light Gray, 23 = Dark Gray
                '''
                pattern = xlwt.Pattern()
                pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                # 背景色是传入数字，这里传入MyColor 转list的下标
                pattern.pattern_fore_colour = list(MyColor).index(color)
                style.pattern = pattern
                # 写入值
                self.sheet_write.write(r=row, c=column, label=value, style=style)
        else:
            print('无效单元格')

    def __write_xlsx(self, row, column, value, color=MyColor.WHITE):
        """
        
        :param row: 要写入的行
        :param column: 要写入的列
        :param value: 要写入的值
        :param color: MyColor.xxxx 不需要.value
        :return:
        """

        row = int(row) + 1
        column = int(column) + 1
        if row > 0 and column > 0:
            cell = self.sheet.cell(row=row, column=column, value=value)
            font = Font(name='Arial',
                        size=11,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color=MyColor.BlACK.value)
            cell.font = font
            patternFill = PatternFill(fill_type='solid',
                                      # start_color='0000FF',
                                      fgColor=color.value)
            cell.fill = patternFill
        else:
            print('无效单元格')


if __name__ == '__main__':
    excel = Excel()
    excel.open_excel(r'C:\Users\yangdang\Desktop\123.xls')
    # excel.open_excel(r'C:\Users\yangdang\Desktop\123.xlsx')
    sheet_name = excel.get_sheets()
    # print(sheet_name)
    li = excel.read_line()
    # print('li', li)
    excel.set_sheet(sheet_name[0])
    excel.write(3, 4, '解决1123', MyColor.BLUE)
    excel.write(1, 1, 'yd', MyColor.BlACK)
    excel.write(7, 7, '解决123123', MyColor.RED)
    excel.write(0, 0, '解决', MyColor.GREEN)
    excel.save()
